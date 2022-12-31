from django.shortcuts import render, redirect
from django.contrib.auth.models import User
import uuid
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from .forms import *
from .models import *
import uuid
from presence_app.settings import EMAIL_HOST_USER, APP_HOST
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import PasswordChangeView
from django.urls import reverse_lazy
from django.http import HttpResponse
from .methods import *
import openpyxl
import json
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Protection
import functools

# ----------------------- Views accounts ------------------- #
# Decorator admin required
def user_admin_required(view_func, redirect_url="presence:index"):
    
    @functools.wraps(view_func)
    def is_admin(request, *args, **kwargs):
        if request.user.user_account.is_admin:
            return view_func(request, *args, **kwargs)
        messages.error(request, "Vous ne pouvez accéder à l'interface demandée")
        return redirect(redirect_url)
    return is_admin


def se_connecter(request):
    try:
        if request.user.is_authenticated:
            return redirect('presence:index')
        
        if request.method == 'POST':
            username = request.POST.get('username')
            password = request.POST.get('password')
            try:
                user = User.objects.get(username=username)
                if user.is_active:
                    user_auth = authenticate(request, username=username, password=password)
                    if user_auth:
                        login(request, user_auth)
                        return redirect('presence:index')     
                    else:
                        messages.error(request, "Vos identifiants sont incorrectes. Réessayer...")
                        return redirect('presence:login')
                else:
                    messages.error(request, "Votre compte est inactif, veuillez contacter l'admin de votre boutique pour la réactivation de votre compte")
                    return redirect('presence:login')
            except Exception as e:
                messages.error(request, "Vos identifiants sont incorrectes. Réessayer...")
                return redirect('presence:login')
        return render(request, 'forms/login.html')
        
    except Exception as e:
        error_html = f"<html><body><h2>Error system: {e}</h2></body></html>"
        return HttpResponse(error_html)

@login_required
@user_admin_required
def utilisateurs_liste(request):
    try:
        users = Enseignant.objects.all().order_by('nom')
        context = {
            'users': users
        }
        return render(request, 'vues/utilisateurs_liste.html', context)
    except Exception as e:
        error_html = f"<html><body><h2>Error system: {e}</h2></body></html>"
        return HttpResponse(error_html)

@login_required
@user_admin_required
def register_user(request):
    try:
        if request.method == 'POST':
            enseignant_form = EnseignantForm(request.POST)
            if enseignant_form.is_valid():
                enseignant = enseignant_form.save(commit=False)
                username = generate_username()
                password = generate_password()
                user = User(username=username)
                user.set_password(password)
                user.save()
                enseignant.user_account = user
                
                subject = "Paramètres de connexion"
                message = f"Bonjour Mr {enseignant.nom} {enseignant.post_nom}, vos paramètres de connexion au système de gestion de presence sont [Nom utilisateur: {username} et Mot de passe: {password}]"
                if enseignant.genre == 'Féminin':
                    message = f"Bonjour Mme {enseignant.nom} {enseignant.post_nom}, vos paramètres de connexion au système de gestion de presence sont [Nom utilisateur: {username} et Mot de passe: {password}]"
                email = enseignant.email
                send_mail_function(subject, message, email)
                
                enseignant.save()
                messages.success(request, f"L'utilisateur {enseignant.nom} {enseignant.post_nom} a été ajouté avec succès")
                return redirect('presence:utilisateurs_liste')
            
        else:
            enseignant_form = EnseignantForm()

        context = {
            'enseignant_form': enseignant_form,
        }    
        return render(request, 'forms/register_user.html', context)
    except Exception as e:
        error_html = f"<html><body><h2>Error system: {e}</h2></body></html>"
        return HttpResponse(error_html)
    
@login_required
@user_admin_required
def update_user(request, id):
    try:
        user = Enseignant.objects.get(id=id)
        if request.method == 'POST':
            enseignant_form = EnseignantForm(request.POST, instance=user)
            if enseignant_form.is_valid():
                enseignant_form.save()
                messages.success(request, f"Les informations de l'utilisateur {user.nom} {user.post_nom} ont été mise à jour avec succès")
                return redirect('presence:utilisateurs_liste')
            
        else:
            enseignant_form = EnseignantForm(instance=user)

        context = {
            'enseignant_form': enseignant_form,
            'user': user,
        }    
        return render(request, 'forms/update_user.html', context)
    except Exception as e:
        error_html = f"<html><body><h2>Error system: {e}</h2></body></html>"
        return HttpResponse(error_html)
    
@login_required
@user_admin_required
def bloque_user(request, id):
    try:
        enseignant = Enseignant.objects.get(id=id)
        user_account = User.objects.get(id=enseignant.user_account.id)
        
        if user_account.is_active:
            user_account.is_active = False
            user_account.save()
            messages.success(request, f"L'utilisateur {enseignant.nom} {enseignant.post_nom} a été bloqué avec succès")
            return redirect('presence:utilisateurs_liste')
        user_account.is_active = True
        user_account.save()
        messages.success(request, f"L'utilisateur {enseignant.nom} {enseignant.post_nom} a été debloqué avec succès")
        return redirect('presence:utilisateurs_liste')
    except Exception as e:
        error_html = f"<html><body><h2>Error system: {e}</h2></body></html>"
        return HttpResponse(error_html)

@login_required
def my_profil(request):
    try:
        return render(request, 'vues/my_profil.html')
    except Exception as e:
        error_html = f"<html><body><h2>Error system: {e}</h2></body></html>"
        return HttpResponse(error_html)
    
@login_required
def update_username(request):
    try:
        user_connected = User.objects.get(id=request.user.id)
        if request.method == 'POST':
            user_form = UserFormUpdate(request.POST, instance=user_connected)
            if user_form.is_valid():
                user_form.save()
                messages.success(request, f"Votre nom d'utilisateur a été mise à jour avec succès")
                return redirect('presence:my_profil')
        else:
            user_form = UserFormUpdate(instance=user_connected)

        context = {
            'user_form': user_form,
        }    
        return render(request, 'forms/update_username.html', context)
    except Exception as e:
        error_html = f"<html><body><h2>Error system: {e}</h2></body></html>"
        return HttpResponse(error_html)

class PasswordChanginView(PasswordChangeView): 
    form_class = PasswwordChangingForm
    success_url = reverse_lazy('presence:message_success_update_password')

@login_required
def message_success_update_password(request):
    try:
        messages.success(request, "Votre mot de passe a été mise à jour avec succès")
        return redirect('presence:my_profil')
    except Exception as e:
        error_html = f"<html><body><h2>Error system: {e}</h2></body></html>"
        return HttpResponse(error_html)
    
def forget_password(request):
    try:
        if request.method == 'POST':
            email = request.POST.get('email')
            try:
                enseignant = Enseignant.objects.get(email=email)
                user_account = User.objects.get(id=enseignant.user_account.id)
                new_password = generate_password()
                user_account.set_password(new_password)
                subject = "Récupéreration du mot de passe"
                message = f"Bonjour Mr {enseignant.nom} {enseignant.post_nom}, votre mot de passe a été reinitialisé avec succès, votre nouveau mot de passe est {new_password}"
                if enseignant.genre == 'Féminin':
                    message = f"Bonjour Mme {enseignant.nom} {enseignant.post_nom}, votre mot de passe a été reinitialisé avec succès, votre nouveau mot de passe est {new_password}"
                send_mail_function(subject, message, email)
                user_account.save()
                messages.success(request, f"Nous vous avons envoyé un nouveau mot de passe sur l'adresse {email}, veuillez consulter vos emails pour récupérer votre compte")
                return redirect('presence:login')
            except Exception as e:
                messages.error(request, f"Aucun utilisateur avec l'adresse email {email} trouvé dans le système")
                return redirect('presence:login')
    except Exception as e:
        error_html = f"<html><body><h2>Error system: {e}</h2></body></html>"
        return HttpResponse(error_html)
            

# ----------------------- End Views accounts ------------------- #

# ----------------------- Views presence management ---------------------- #

@login_required
def index(request):
    try:
        enseignant = request.user.user_account
        mes_sceances = Sceance.objects.filter(enseignant=enseignant).order_by('-id')
        context = {
            'mes_sceances': mes_sceances,
        }
        return render(request, 'vues/index.html', context)
    except Exception as e:
        error_html = f"<html><body><h2>Error system: {e}</h2></body></html>"
        return HttpResponse(error_html)

@login_required
def add_sceance(request):
    try:
        if request.method == 'POST':
            sceance_form = SceanceForm(request.POST)
            if sceance_form.is_valid():
                sceance = sceance_form.save(commit=False)
                sceance.enseignant = request.user.user_account
                sceance.save()
                messages.success(request, f"La scéance {sceance.cours} de la {sceance.promotion} a été créée avec succès")
                return redirect('presence:index')
        else:
            sceance_form = SceanceForm()

        context = {
            'sceance_form': sceance_form,
        }    
        return render(request, 'forms/add_sceance.html', context)
    except Exception as e:
        error_html = f"<html><body><h2>Error system: {e}</h2></body></html>"
        return HttpResponse(error_html)
    
@login_required
def update_sceance(request, id):
    try:
        sceance = Sceance.objects.get(id=id)
        if request.method == 'POST':
            sceance_form = SceanceForm(request.POST, instance=sceance)
            if sceance_form.is_valid():
                sceance_form.save()
                messages.success(request, f"La mise à jour de la scéance {sceance.cours} de la {sceance.promotion} a été effectuée avec succès")
                return redirect('presence:index')
        else:
            sceance_form = SceanceForm(instance=sceance)

        context = {
            'sceance_form': sceance_form,
            'sceance': sceance,
        }    
        return render(request, 'forms/update_sceance.html', context)
    except Exception as e:
        error_html = f"<html><body><h2>Error system: {e}</h2></body></html>"
        return HttpResponse(error_html)

def verify_create_seance_day(sceance):
    try:
        sceances = SceanceDay.objects.filter(sceance=sceance)
        sceance_day = sceances.get(date_sceance=date.today())
        return sceance_day
    except Exception as e:
        sceance_day = SceanceDay.objects.create(sceance=sceance)
        sceance_day.save()
        return sceance_day

@login_required
def open_sceance(request, id):
    try:
        sceance = Sceance.objects.get(id=id)
        if request.method =="POST":
            participant_id = request.POST.get("content_id")
            participant = Participant.objects.get(id=participant_id)
            sceance_day = verify_create_seance_day(sceance)
            
            if participant.presences.filter(id=sceance_day.id):
                participant.presences.remove(sceance_day)
                present=False
            else:
                participant.presences.add(sceance_day)
                present=True
            context = {
                "present": present,
                }
            return HttpResponse(json.dumps(context), content_type='application/json')
        participants = Participant.objects.filter(sceance=sceance).order_by('nom', 'post_nom', 'pre_nom')
        context = {
            'sceance': sceance,
            'participants': participants,
        }
        return render(request, 'vues/open_sceance.html', context)
    except Exception as e:
        error_html = f"<html><body><h2>Error system: {e}</h2></body></html>"
        return HttpResponse(error_html)

@login_required
def import_data_excel(request, id):
    try:
        sceance = Sceance.objects.get(id=id)
        if request.method == 'POST' and request.FILES['myfile']:
            excel_file = request.FILES["myfile"]


            wb = openpyxl.load_workbook(excel_file)

            # getting a particular sheet by name out of many sheets
            worksheet = wb["Sheet1"]
            print(worksheet)

            excel_data = list()
            for row in worksheet.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data.append(row_data)

            date_filter = excel_data[1:]
            for data in date_filter:
                mat = int(data[0])
                nom = data[1]
                post_nom = data[2]
                pre_nom = data[3]
                genre = data[4]
                
                # Test existance
                participant_is_exist = Participant.objects.filter(sceance=sceance).filter(matricule=mat)
                if participant_is_exist.exists():
                    print(True)
                    participant = participant_is_exist.latest('id')
                    print(participant)
                    participant.nom = nom
                    participant.post_nom = post_nom
                    participant.pre_nom = pre_nom
                    participant.genre = genre
                    participant.save()
                else:
                    participant = Participant.objects.create(
                        matricule = mat,
                        nom = nom, 
                        post_nom = post_nom,
                        pre_nom = pre_nom,
                        genre = genre,
                        sceance=sceance,
                        )
                    participant.save()
                    print(False)
                    
            
            messages.success(request, f"Le chargement des participants s'est effectuée avec succès")
            return redirect('presence:open_sceance', sceance.id)
        
        messages.error(request, f"La methode que vous voulez utiliser n'est pas post")
        return redirect('presence:open_sceance')
    except Exception as e:
        error_html = f"<html><body><h2>Error system: {e}</h2></body></html>"
        return HttpResponse(error_html)

@login_required
def export_data_to_excel(request, id):
    sceance = Sceance.objects.get(id=id)
    participants = Participant.objects.filter(sceance=sceance).order_by('nom', 'post_nom')
        
    excelfile = BytesIO()
    workbook = Workbook()
    workbook.remove(workbook.active)
    worksheet = workbook.create_sheet(title='Liste de presence', index=1)
    output_filename = f'liste_presence_{sceance.cours}_{sceance.promotion}.xlsx'

    columns = ['Mat', 'Nom', 'Postnom', 'Prénom', 'Genre', 'Percent']
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(bold=True)

    for _, participant in enumerate(participants, 1):
        row_num += 1

        row = [
            participant.matricule,
            participant.nom,
            participant.post_nom,
            participant.pre_nom,
            participant.genre,
            participant.pourcentage_presence(),
        ]


        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.protection = Protection(locked=True)
    workbook.save(excelfile)
    
    http_response = HttpResponse(
            excelfile.getvalue(),
            content_type='application/vnd.ms-excel'
        )
    http_response['Content-Disposition'] = f'filename={output_filename}'
    return http_response

@login_required
@user_admin_required
def admin_views(request):
    try:
        sceances = Sceance.objects.all().order_by('-id')
        context = {
            'sceances': sceances,
        }
        return render(request, 'vues/admin_views.html', context)
    except Exception as e:
        error_html = f"<html><body><h2>Error system: {e}</h2></body></html>"
        return HttpResponse(error_html)
    
@login_required
@user_admin_required
def detail_sceance(request, id):
    try:
        sceance = Sceance.objects.get(id=id)
        participants = Participant.objects.filter(sceance=sceance).order_by('nom', 'post_nom', 'pre_nom')
        context = {
            'sceance': sceance,
            'participants': participants,
        }
        return render(request, 'vues/detail_sceance.html', context)
    except Exception as e:
        error_html = f"<html><body><h2>Error system: {e}</h2></body></html>"
        return HttpResponse(error_html)
# ----------------------- End Views presence management ---------------------- #


    